"""
Seed generation and simulation type router.
Handles file uploads, URL extraction, search-based seed generation,
and simulation type definitions.
"""

import io
import json
import os
from typing import Optional

from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from pydantic import BaseModel

from services.seed_factory import (
    get_all_types,
    get_sub_templates,
    load_type_definition,
    get_canonical_seed,
    apply_type_defaults,
    _get_requirements,
)
from services.seed_validator import validate_seed
from services.tavily_service import search as tavily_search, extract as tavily_extract
from services.llm_service import extract_seed_json

router = APIRouter(tags=["seeds"])

PROMPTS_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "prompts")


def _load_extraction_prompt() -> str:
    """Load the seed extraction prompt template."""
    path = os.path.join(PROMPTS_DIR, "seed_extraction.txt")
    with open(path, "r") as f:
        return f.read()


def _build_extraction_prompt(
    extracted_text: str,
    simulation_type_id: str = "corporate_strategy",
) -> str:
    """Build the full extraction prompt with type-specific instructions."""
    template = _load_extraction_prompt()
    type_def = load_type_definition(simulation_type_id)

    type_name = type_def["name"] if type_def else "Custom Simulation"

    # Build type-specific extraction instructions
    instructions = ""
    if type_def:
        # Check for extraction_directive from existing types JSON
        extraction_directive = type_def.get("extraction_directive", "")
        if extraction_directive:
            instructions = extraction_directive
        else:
            reqs = _get_requirements(type_def)
            roles = reqs.get("required_roles", [])
            actions = reqs.get("required_actions", [])
            env_vars = reqs.get("required_environment", [])

            if roles:
                instructions += f"Required roles to identify: {', '.join(roles)}\n"
            if actions:
                instructions += f"Each actor should have actions including: {', '.join(actions)}\n"
            if env_vars:
                instructions += f"Environment must include: {', '.join(env_vars)}\n"

    extraction_focus = ""
    if type_def:
        extraction_focus = type_def.get("description", type_def.get("tagline", ""))

    return template.format(
        simulation_type_name=type_name,
        simulation_type_id=simulation_type_id,
        type_specific_extraction_instructions=instructions or "No specific directives.",
        extraction_focus=extraction_focus or "general strategic analysis",
        extracted_text=extracted_text,
    )


def _extract_text_from_pdf(content: bytes) -> str:
    """Extract text from a PDF file."""
    import pdfplumber

    text_parts = []
    with pdfplumber.open(io.BytesIO(content)) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text_parts.append(page_text)
    return "\n\n".join(text_parts)


def _extract_text_from_docx(content: bytes) -> str:
    """Extract text from a DOCX file."""
    from docx import Document

    doc = Document(io.BytesIO(content))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n\n".join(paragraphs)


def _extract_text_from_xlsx(content: bytes) -> str:
    """Extract text from an XLSX file."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    text_parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        text_parts.append(f"=== Sheet: {sheet_name} ===")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                text_parts.append("\t".join(cells))
    return "\n".join(text_parts)


def _extract_text_from_txt(content: bytes) -> str:
    """Extract text from a plain text file."""
    for encoding in ("utf-8", "latin-1", "cp1252"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


# ========================
# File Upload Seed Generation
# ========================

@router.post("/api/seed/from-file")
async def seed_from_file(
    file: UploadFile = File(...),
    simulation_type: str = Form("corporate_strategy"),
):
    """
    Upload a PDF/DOCX/TXT/XLSX file, extract text, and generate a seed scaffold.
    Returns the extracted text and a prompt ready for LLM processing.
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename provided")

    content = await file.read()
    filename_lower = file.filename.lower()

    # Extract text based on file type
    if filename_lower.endswith(".pdf"):
        extracted_text = _extract_text_from_pdf(content)
    elif filename_lower.endswith((".docx", ".doc")):
        extracted_text = _extract_text_from_docx(content)
    elif filename_lower.endswith((".xlsx", ".xls")):
        extracted_text = _extract_text_from_xlsx(content)
    elif filename_lower.endswith((".txt", ".md", ".csv")):
        extracted_text = _extract_text_from_txt(content)
    else:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file type: {file.filename}. Supported: PDF, DOCX, TXT, XLSX, MD, CSV",
        )

    if not extracted_text.strip():
        raise HTTPException(status_code=400, detail="No text could be extracted from the file")

    # Build the extraction prompt and call LLM to generate seed
    prompt = _build_extraction_prompt(extracted_text[:15000], simulation_type)

    try:
        generated_seed = await extract_seed_json(prompt)
    except Exception as e:
        # Fallback to scaffold if LLM fails
        generated_seed = get_canonical_seed(simulation_type) or {}
        generated_seed["_llm_error"] = str(e)

    return {
        "success": True,
        "data": {
            "extracted_text": extracted_text[:500],
            "text_length": len(extracted_text),
            "seed": generated_seed,
            "seed_scaffold": generated_seed,
            "simulation_type": simulation_type,
            "filename": file.filename,
        },
    }


# ========================
# URL-based Seed Generation
# ========================

class UrlSeedRequest(BaseModel):
    url: str
    simulation_type: str = "corporate_strategy"


@router.post("/api/seed/from-url")
async def seed_from_url(body: UrlSeedRequest):
    """
    Extract content from a URL using Tavily, then generate a seed scaffold.
    """
    result = await tavily_extract(body.url)

    # Tavily extract returns { results: [ { url, raw_content } ] }
    results = result.get("results", [])
    if not results:
        raise HTTPException(
            status_code=400,
            detail=f"Could not extract content from URL: {body.url}",
        )

    extracted_text = results[0].get("raw_content", "")
    if not extracted_text:
        extracted_text = results[0].get("content", "")

    if not extracted_text.strip():
        raise HTTPException(status_code=400, detail="No text could be extracted from the URL")

    prompt = _build_extraction_prompt(extracted_text[:15000], body.simulation_type)

    try:
        generated_seed = await extract_seed_json(prompt)
    except Exception as e:
        generated_seed = get_canonical_seed(body.simulation_type) or {}
        generated_seed["_llm_error"] = str(e)

    return {
        "success": True,
        "data": {
            "extracted_text": extracted_text[:500],
            "text_length": len(extracted_text),
            "seed": generated_seed,
            "seed_scaffold": generated_seed,
            "simulation_type": body.simulation_type,
            "source_url": body.url,
        },
    }


# ========================
# Search-based Seed Generation
# ========================

class SearchSeedRequest(BaseModel):
    query: str
    simulation_type: str = "corporate_strategy"
    max_results: int = 10


@router.post("/api/seed/from-search")
async def seed_from_search(body: SearchSeedRequest):
    """
    Search the web using Tavily, aggregate results, and generate a seed scaffold.
    """
    result = await tavily_search(body.query, max_results=body.max_results)

    answer = result.get("answer", "")
    search_results = result.get("results", [])

    if not search_results and not answer:
        raise HTTPException(
            status_code=400,
            detail=f"No results found for query: {body.query}",
        )

    # Aggregate text from search results
    text_parts = []
    if answer:
        text_parts.append(f"SUMMARY:\n{answer}\n")

    sources = []
    for r in search_results:
        title = r.get("title", "")
        content = r.get("raw_content") or r.get("content", "")
        url = r.get("url", "")
        if content:
            text_parts.append(f"SOURCE: {title}\nURL: {url}\n{content}\n")
            sources.append({"title": title, "url": url})

    extracted_text = "\n---\n".join(text_parts)
    prompt = _build_extraction_prompt(extracted_text[:15000], body.simulation_type)

    try:
        generated_seed = await extract_seed_json(prompt)
    except Exception as e:
        generated_seed = get_canonical_seed(body.simulation_type) or {}
        generated_seed["_llm_error"] = str(e)

    return {
        "success": True,
        "data": {
            "extracted_text": extracted_text[:500],
            "text_length": len(extracted_text),
            "seed": generated_seed,
            "seed_scaffold": generated_seed,
            "simulation_type": body.simulation_type,
            "search_query": body.query,
            "sources": sources,
            "answer": answer,
        },
    }


# ========================
# Seed Validation
# ========================

class ValidateSeedRequest(BaseModel):
    seed: dict
    simulation_type: str = "corporate_strategy"


@router.post("/api/seed/validate")
async def validate_seed_endpoint(body: ValidateSeedRequest):
    """Validate a seed document against simulation type requirements."""
    result = validate_seed(body.seed, body.simulation_type)
    return {"success": True, "data": result}


# ========================
# Simulation Types
# ========================

@router.get("/api/simulation-types")
async def list_simulation_types():
    """Return all simulation type definitions."""
    types = get_all_types()
    return {"success": True, "data": types}


@router.get("/api/simulation-types/{type_id}")
async def get_simulation_type(type_id: str):
    """Return a single simulation type definition."""
    type_def = load_type_definition(type_id)
    if not type_def:
        raise HTTPException(status_code=404, detail=f"Simulation type not found: {type_id}")
    return {"success": True, "data": type_def}


@router.get("/api/simulation-types/{type_id}/templates")
async def list_sub_templates(type_id: str):
    """Return sub-templates for a simulation type."""
    type_def = load_type_definition(type_id)
    if not type_def:
        raise HTTPException(status_code=404, detail=f"Simulation type not found: {type_id}")

    templates = get_sub_templates(type_id)
    return {"success": True, "data": templates}


@router.get("/api/simulation-types/{type_id}/scaffold")
async def get_seed_scaffold(type_id: str):
    """Return the canonical seed scaffold for a simulation type."""
    scaffold = get_canonical_seed(type_id)
    if not scaffold:
        raise HTTPException(status_code=404, detail=f"Simulation type not found: {type_id}")
    return {"success": True, "data": scaffold}
